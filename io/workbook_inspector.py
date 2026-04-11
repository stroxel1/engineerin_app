"""Workbook inspection helpers.

These functions are intentionally lightweight so they can be used even before
full import normalization is implemented.
"""

from __future__ import annotations

from hashlib import sha256
from pathlib import Path
from typing import Any

from engineering_app.io.contracts import SheetPreview, WorkbookSource


try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class WorkbookInspectorError(RuntimeError):
    pass


def _coerce_cell(value: Any) -> Any:
    if isinstance(value, (int, float, str, bool)) or value is None:
        return value
    return str(value)


def _file_checksum(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sheet_stats(rows: list[list[Any]]) -> dict[str, Any]:
    non_empty_rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    flat_cells = [cell for row in rows for cell in row if cell not in (None, "")]
    numeric_cells = [cell for cell in flat_cells if isinstance(cell, (int, float)) and not isinstance(cell, bool)]
    text_cells = [cell for cell in flat_cells if isinstance(cell, str) and cell.strip()]

    return {
        "sample_non_empty_row_count": len(non_empty_rows),
        "sample_populated_cell_count": len(flat_cells),
        "sample_numeric_cell_count": len(numeric_cells),
        "sample_text_cell_count": len(text_cells),
        "sample_numeric_fraction": (len(numeric_cells) / len(flat_cells)) if flat_cells else 0.0,
    }


def _header_candidates(rows: list[list[Any]], limit: int = 3) -> list[list[str]]:
    candidates: list[list[str]] = []
    for row in rows:
        cleaned = [str(cell).strip() for cell in row if cell not in (None, "")]
        if len(cleaned) >= 2:
            candidates.append(cleaned)
        if len(candidates) >= limit:
            break
    return candidates


def get_workbook_source(path: str | Path) -> WorkbookSource:
    file_path = Path(path)
    stat = file_path.stat()
    return WorkbookSource(
        path=str(file_path),
        modified_time=stat.st_mtime,
        checksum=_file_checksum(file_path),
    )


def inspect_workbook(path: str | Path, sample_rows: int = 10) -> dict[str, Any]:
    if load_workbook is None:
        raise WorkbookInspectorError("openpyxl is not installed")

    file_path = Path(path)
    workbook = load_workbook(file_path, data_only=False)
    previews: list[SheetPreview] = []

    for sheet in workbook.worksheets:
        rows: list[list[Any]] = []
        formula_cells = 0
        for row in sheet.iter_rows(min_row=1, max_row=min(sample_rows, sheet.max_row)):
            cooked_row = []
            for cell in row:
                value = _coerce_cell(cell.value)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
                cooked_row.append(value)
            rows.append(cooked_row)

        stats = _sheet_stats(rows)
        previews.append(
            SheetPreview(
                sheet_name=sheet.title,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                sample_rows=rows,
                merged_ranges=[str(rng) for rng in sheet.merged_cells.ranges],
                hidden=sheet.sheet_state != "visible",
                freeze_panes=str(sheet.freeze_panes) if sheet.freeze_panes else None,
                formula_cell_count=formula_cells,
                header_candidates=_header_candidates(rows),
                stats=stats,
            )
        )

    return {
        "source": get_workbook_source(file_path).__dict__,
        "sheet_previews": [preview.__dict__ for preview in previews],
    }
